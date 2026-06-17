"""CAT Habit Tracker — FastAPI backend on Supabase Postgres (SQLAlchemy async)."""
from dotenv import load_dotenv
from pathlib import Path
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

import os
import io
import csv
import json
import uuid
import logging
import bcrypt
import httpx
from datetime import datetime, timezone, timedelta
from typing import Optional, List

from fastapi import FastAPI, APIRouter, HTTPException, Request, Response, UploadFile, File, Form, Depends
from starlette.middleware.cors import CORSMiddleware
from pydantic import BaseModel, EmailStr, Field
from openpyxl import load_workbook
from sqlalchemy import select, delete, update, func
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.dialects.postgresql import insert as pg_insert

from database import AsyncSessionLocal, engine
from models import Base, User, UserSession, Habit, TimetableEntry

app = FastAPI()
api_router = APIRouter(prefix="/api")

EMERGENT_AUTH_URL = "https://demobackend.emergentagent.com/auth/v1/env/oauth/session-data"
SESSION_DAYS = 7
COOKIE_KW = dict(httponly=True, secure=True, samesite="none", path="/")


# ─── DB Dependency ─────────────────────────────────────────────────
async def get_db() -> AsyncSession:
    async with AsyncSessionLocal() as s:
        try:
            yield s
        finally:
            await s.close()


# ─── Pydantic payloads ─────────────────────────────────────────────
class RegisterPayload(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)


class LoginPayload(BaseModel):
    email: EmailStr
    password: str


class SessionPayload(BaseModel):
    session_id: str


class SettingsPayload(BaseModel):
    exam_name: Optional[str] = None
    target_exam_date: Optional[str] = None


class HabitCreate(BaseModel):
    title: str = Field(min_length=1)
    icon: Optional[str] = "✦"
    category: Optional[str] = "Custom"


class TimetableIn(BaseModel):
    subject: str
    teacher: Optional[str] = ""
    start_time: str
    end_time: Optional[str] = ""
    date: str
    duration: Optional[int] = 60
    notes: Optional[str] = ""


# ─── Helpers ───────────────────────────────────────────────────────
def hash_password(p: str) -> str:
    return bcrypt.hashpw(p.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def verify_password(p: str, h: str) -> bool:
    try:
        return bcrypt.checkpw(p.encode("utf-8"), h.encode("utf-8"))
    except Exception:
        return False


def now_utc() -> datetime:
    return datetime.now(timezone.utc)


def today_str() -> str:
    return now_utc().strftime("%Y-%m-%d")


def public_user(u: User) -> dict:
    return {
        "user_id": u.user_id,
        "email": u.email,
        "name": u.name,
        "picture": u.picture,
        "exam_name": u.exam_name,
        "target_exam_date": u.target_exam_date,
        "role": u.role,
    }


async def issue_session(s: AsyncSession, response: Response, user_id: str) -> str:
    token = uuid.uuid4().hex + uuid.uuid4().hex
    expires = now_utc() + timedelta(days=SESSION_DAYS)
    s.add(UserSession(session_token=token, user_id=user_id, expires_at=expires, created_at=now_utc()))
    await s.commit()
    response.set_cookie(
        key="session_token", value=token,
        max_age=SESSION_DAYS * 24 * 60 * 60,
        **COOKIE_KW,
    )
    return token


DEFAULT_HABITS_TEMPLATE = [
    {"title": "Quant: 25 LOD questions", "category": "QA", "icon": "🧮"},
    {"title": "VARC: 2 RC passages", "category": "VARC", "icon": "📚"},
    {"title": "LRDI: 1 caselet set", "category": "LRDI", "icon": "🧩"},
    {"title": "Mock analysis (30 min)", "category": "Mock", "icon": "📊"},
    {"title": "Vocab: 10 new words", "category": "VARC", "icon": "🔤"},
]


async def seed_user_defaults(s: AsyncSession, user_id: str):
    cnt = await s.scalar(select(func.count()).select_from(Habit).where(Habit.user_id == user_id))
    if cnt and cnt > 0:
        return
    for h in DEFAULT_HABITS_TEMPLATE:
        s.add(Habit(
            habit_id=f"h_{uuid.uuid4().hex[:10]}",
            user_id=user_id, title=h["title"], category=h["category"], icon=h["icon"],
            completed_dates=[], created_at=now_utc(),
        ))
    await s.commit()


async def get_current_user(request: Request, db: AsyncSession = Depends(get_db)) -> User:
    token = request.cookies.get("session_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")

    session = await db.scalar(select(UserSession).where(UserSession.session_token == token))
    if not session:
        raise HTTPException(status_code=401, detail="Invalid session")
    if session.expires_at < now_utc():
        raise HTTPException(status_code=401, detail="Session expired")

    user = await db.scalar(select(User).where(User.user_id == session.user_id))
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user


# ─── Auth ──────────────────────────────────────────────────────────
@api_router.get("/")
async def root():
    return {"message": "CAT Habit Tracker API (Supabase)"}


@api_router.post("/auth/register")
async def register(payload: RegisterPayload, response: Response, db: AsyncSession = Depends(get_db)):
    email = payload.email.lower().strip()
    existing = await db.scalar(select(User).where(User.email == email))
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    u = User(
        user_id=f"user_{uuid.uuid4().hex[:12]}",
        email=email, name=payload.name.strip(),
        password_hash=hash_password(payload.password),
        picture=None, exam_name="CAT",
        target_exam_date="2026-11-30", role="user",
        created_at=now_utc(),
    )
    db.add(u)
    await db.commit()
    await db.refresh(u)
    await seed_user_defaults(db, u.user_id)
    await issue_session(db, response, u.user_id)
    return public_user(u)


@api_router.post("/auth/login")
async def login(payload: LoginPayload, response: Response, db: AsyncSession = Depends(get_db)):
    email = payload.email.lower().strip()
    u = await db.scalar(select(User).where(User.email == email))
    if not u or not u.password_hash or not verify_password(payload.password, u.password_hash):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    await seed_user_defaults(db, u.user_id)
    await issue_session(db, response, u.user_id)
    return public_user(u)


@api_router.post("/auth/session")
async def emergent_session(payload: SessionPayload, response: Response, db: AsyncSession = Depends(get_db)):
    async with httpx.AsyncClient(timeout=15.0) as http:
        r = await http.get(EMERGENT_AUTH_URL, headers={"X-Session-ID": payload.session_id})
    if r.status_code != 200:
        raise HTTPException(status_code=401, detail="Invalid session")
    data = r.json()
    email = data["email"].lower().strip()
    u = await db.scalar(select(User).where(User.email == email))
    if u:
        u.name = data["name"]
        u.picture = data.get("picture")
    else:
        u = User(
            user_id=f"user_{uuid.uuid4().hex[:12]}",
            email=email, name=data["name"], picture=data.get("picture"),
            password_hash=None, exam_name="CAT",
            target_exam_date="2026-11-30", role="user",
            created_at=now_utc(),
        )
        db.add(u)
    await db.commit()
    await db.refresh(u)
    await seed_user_defaults(db, u.user_id)
    await issue_session(db, response, u.user_id)
    return public_user(u)


@api_router.get("/auth/me")
async def auth_me(user: User = Depends(get_current_user)):
    return public_user(user)


@api_router.post("/auth/logout")
async def logout(request: Request, response: Response, db: AsyncSession = Depends(get_db)):
    token = request.cookies.get("session_token")
    if token:
        await db.execute(delete(UserSession).where(UserSession.session_token == token))
        await db.commit()
    response.delete_cookie(key="session_token", path="/", samesite="none", secure=True)
    return {"ok": True}


# ─── Settings ──────────────────────────────────────────────────────
@api_router.get("/me/settings")
async def get_settings(user: User = Depends(get_current_user)):
    return {"exam_name": user.exam_name, "target_exam_date": user.target_exam_date}


@api_router.put("/me/settings")
async def update_settings(payload: SettingsPayload, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if payload.exam_name is not None:
        user.exam_name = (payload.exam_name.strip()[:50]) or "CAT"
    if payload.target_exam_date is not None:
        try:
            datetime.strptime(payload.target_exam_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="Date must be YYYY-MM-DD")
        user.target_exam_date = payload.target_exam_date
    db.add(user)
    await db.commit()
    return {"exam_name": user.exam_name, "target_exam_date": user.target_exam_date}


# ─── Habits ────────────────────────────────────────────────────────
def habit_dict(h: Habit, today: str) -> dict:
    return {
        "habit_id": h.habit_id, "user_id": h.user_id, "title": h.title,
        "category": h.category, "icon": h.icon,
        "completed_dates": h.completed_dates or [],
        "created_at": h.created_at.isoformat() if h.created_at else None,
        "completed_today": today in (h.completed_dates or []),
    }


@api_router.get("/habits")
async def list_habits(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(Habit).where(Habit.user_id == user.user_id).order_by(Habit.created_at))
    today = today_str()
    return [habit_dict(h, today) for h in r.scalars().all()]


@api_router.post("/habits")
async def create_habit(payload: HabitCreate, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    h = Habit(
        habit_id=f"h_{uuid.uuid4().hex[:10]}", user_id=user.user_id,
        title=payload.title.strip()[:120],
        category=(payload.category or "Custom")[:30],
        icon=(payload.icon or "✦")[:4],
        completed_dates=[], created_at=now_utc(),
    )
    db.add(h)
    await db.commit()
    await db.refresh(h)
    return habit_dict(h, today_str())


@api_router.delete("/habits/{habit_id}")
async def delete_habit(habit_id: str, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    await db.execute(delete(Habit).where(Habit.habit_id == habit_id, Habit.user_id == user.user_id))
    await db.commit()
    return {"ok": True}


@api_router.post("/habits/{habit_id}/toggle")
async def toggle_habit(habit_id: str, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    h = await db.scalar(select(Habit).where(Habit.habit_id == habit_id, Habit.user_id == user.user_id))
    if not h:
        raise HTTPException(status_code=404, detail="Habit not found")
    today = today_str()
    dates = set(h.completed_dates or [])
    if today in dates:
        dates.remove(today); completed = False
    else:
        dates.add(today); completed = True
    h.completed_dates = sorted(dates)
    db.add(h)
    await db.commit()
    return {"completed_today": completed, "habit_id": habit_id}


@api_router.get("/habits/stats")
async def habit_stats(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(Habit).where(Habit.user_id == user.user_id))
    habits = r.scalars().all()
    total = len(habits)

    day_counts: dict = {}
    for h in habits:
        for d in h.completed_dates or []:
            day_counts[d] = day_counts.get(d, 0) + 1

    today = now_utc().date()

    # Heatmap (last 30 days)
    heatmap = []
    for i in range(29, -1, -1):
        d = today - timedelta(days=i)
        k = d.isoformat()
        heatmap.append({"date": k, "count": day_counts.get(k, 0), "total": total})

    # Weekly (last 7 days)  — Mon=0..Sun=6
    day_letters = {0: "M", 1: "T", 2: "W", 3: "T", 4: "F", 5: "S", 6: "S"}
    weekly = []
    for i in range(6, -1, -1):
        d = today - timedelta(days=i)
        k = d.isoformat()
        c = day_counts.get(k, 0)
        weekly.append({"date": k, "day": day_letters[d.weekday()], "pct": round((c / total) * 100) if total else 0})

    # Current streak — allow today incomplete
    streak = 0
    cur = today
    for i in range(365):
        k = cur.isoformat()
        c = day_counts.get(k, 0)
        if c == 0:
            if i == 0:
                cur = cur - timedelta(days=1); continue
            break
        streak += 1
        cur = cur - timedelta(days=1)

    # Best streak
    keys = sorted([k for k, v in day_counts.items() if v > 0])
    best = streak
    if keys:
        run = 1
        for i in range(1, len(keys)):
            prev = datetime.fromisoformat(keys[i - 1]).date()
            curr = datetime.fromisoformat(keys[i]).date()
            if (curr - prev).days == 1:
                run += 1; best = max(best, run)
            else:
                run = 1
        best = max(best, streak)

    completed_today = day_counts.get(today.isoformat(), 0)
    pct = round((completed_today / total) * 100) if total else 0
    return {
        "streak": streak, "best_streak": best,
        "completed_today": completed_today, "total_habits": total,
        "completion_pct": pct, "heatmap": heatmap, "weekly": weekly,
    }


# ─── Timetable ─────────────────────────────────────────────────────
def tt_dict(e: TimetableEntry) -> dict:
    return {
        "entry_id": e.entry_id, "user_id": e.user_id,
        "subject": e.subject, "teacher": e.teacher,
        "start_time": e.start_time, "end_time": e.end_time,
        "date": e.date, "duration": e.duration, "notes": e.notes,
        "created_at": e.created_at.isoformat() if e.created_at else None,
    }


@api_router.get("/timetable")
async def list_timetable(date: Optional[str] = None, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    q = select(TimetableEntry).where(TimetableEntry.user_id == user.user_id)
    if date:
        q = q.where(TimetableEntry.date == date)
    q = q.order_by(TimetableEntry.start_time.asc())
    r = await db.execute(q)
    return [tt_dict(e) for e in r.scalars().all()]


@api_router.post("/timetable")
async def create_tt(entry: TimetableIn, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    e = TimetableEntry(
        entry_id=f"t_{uuid.uuid4().hex[:10]}", user_id=user.user_id,
        **entry.model_dump(), created_at=now_utc(),
    )
    db.add(e); await db.commit(); await db.refresh(e)
    return tt_dict(e)


@api_router.delete("/timetable/{entry_id}")
async def delete_tt(entry_id: str, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    await db.execute(delete(TimetableEntry).where(TimetableEntry.entry_id == entry_id, TimetableEntry.user_id == user.user_id))
    await db.commit()
    return {"ok": True}


def _norm_row(row: dict) -> Optional[dict]:
    lower = {str(k).strip().lower(): v for k, v in row.items() if k is not None}
    def pick(*names):
        for n in names:
            v = lower.get(n)
            if v not in (None, ""):
                return str(v).strip()
        return ""
    subject = pick("subject", "topic", "task", "title")
    date = pick("date", "day")
    if not subject or not date:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            date = datetime.strptime(date, fmt).strftime("%Y-%m-%d"); break
        except Exception:
            continue
    return {
        "subject": subject[:120],
        "teacher": pick("teacher", "instructor", "by")[:80],
        "start_time": pick("start_time", "start", "time", "from")[:8],
        "end_time": pick("end_time", "end", "to")[:8],
        "date": date,
        "duration": int(pick("duration") or "60") if (pick("duration") or "").isdigit() else 60,
        "notes": pick("notes", "remarks", "description")[:200],
    }


def _parse_csv(raw: bytes) -> List[dict]:
    text = raw.decode("utf-8-sig", errors="replace")
    return [r for r in (_norm_row(x) for x in csv.DictReader(io.StringIO(text))) if r]


def _parse_json(raw: bytes) -> List[dict]:
    data = json.loads(raw.decode("utf-8-sig", errors="replace"))
    if isinstance(data, dict):
        data = data.get("entries") or data.get("data") or [data]
    if not isinstance(data, list):
        raise HTTPException(status_code=400, detail="JSON must be an array or {entries:[...]}")
    return [r for r in (_norm_row(x) for x in data) if r]


def _parse_xlsx(raw: bytes) -> List[dict]:
    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    out = []
    for r in rows[1:]:
        row = {headers[i]: r[i] for i in range(min(len(headers), len(r)))}
        if isinstance(row.get("date"), datetime):
            row["date"] = row["date"].strftime("%Y-%m-%d")
        for tkey in ("start_time", "end_time", "start", "end", "time"):
            v = row.get(tkey)
            if hasattr(v, "strftime"):
                row[tkey] = v.strftime("%H:%M")
        norm = _norm_row(row)
        if norm:
            out.append(norm)
    return out


@api_router.post("/timetable/upload")
async def upload_tt(
    file: UploadFile = File(...),
    mode: str = Form("add"),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    if mode not in ("add", "replace"):
        raise HTTPException(status_code=400, detail="mode must be 'add' or 'replace'")
    raw = await file.read()
    name = (file.filename or "").lower()
    try:
        if name.endswith(".csv"): entries = _parse_csv(raw)
        elif name.endswith(".json"): entries = _parse_json(raw)
        elif name.endswith((".xlsx", ".xls")): entries = _parse_xlsx(raw)
        else: raise HTTPException(status_code=400, detail="Unsupported file type. Use CSV / XLSX / JSON.")
    except HTTPException: raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Parse failed: {e}")

    if not entries:
        raise HTTPException(status_code=400, detail="No valid rows found in file.")

    if mode == "replace":
        dates = sorted({e["date"] for e in entries})
        await db.execute(delete(TimetableEntry).where(TimetableEntry.user_id == user.user_id, TimetableEntry.date.in_(dates)))

    for e in entries:
        db.add(TimetableEntry(
            entry_id=f"t_{uuid.uuid4().hex[:10]}", user_id=user.user_id,
            **e, created_at=now_utc(),
        ))
    await db.commit()
    return {"inserted": len(entries), "mode": mode, "dates": sorted({e["date"] for e in entries})}


# ─── App setup ─────────────────────────────────────────────────────
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


@app.on_event("startup")
async def startup():
    # Seed admin + test user (idempotent)
    async with AsyncSessionLocal() as s:
        for email_env, pw_env, role, name in [
            ("ADMIN_EMAIL", "ADMIN_PASSWORD", "admin", "Admin"),
            ("TEST_USER_EMAIL", "TEST_USER_PASSWORD", "user", "Aspirant"),
        ]:
            email = os.environ.get(email_env)
            pw = os.environ.get(pw_env)
            if not email or not pw:
                continue
            email = email.lower()
            u = await s.scalar(select(User).where(User.email == email))
            ph = hash_password(pw)
            if not u:
                u = User(
                    user_id=f"user_{uuid.uuid4().hex[:12]}",
                    email=email, name=name, password_hash=ph, picture=None,
                    exam_name="CAT", target_exam_date="2026-11-30",
                    role=role, created_at=now_utc(),
                )
                s.add(u)
                await s.commit()
                await seed_user_defaults(s, u.user_id)
            elif not u.password_hash or not verify_password(pw, u.password_hash):
                u.password_hash = ph
                s.add(u)
                await s.commit()


@app.on_event("shutdown")
async def shutdown():
    await engine.dispose()
